import csv
import json
import io
from django.http import HttpResponse
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from survey.models import Response, Answer
from survey.questions import QUESTIONS


@staff_member_required
def export_csv(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="respondents.csv"'
    response.write('\ufeff')

    writer = csv.writer(response)
    headers = ['ID', 'Ism', 'Yosh', 'Viloyat', 'IP', 'Sana']
    for q in QUESTIONS:
        if q['type'] != 'consent':
            headers.append(f"Savol {q['number']}")
    writer.writerow(headers)

    respondents = Response.objects.all().prefetch_related('answers')
    for r in respondents:
        row = [r.id, r.name or '', r.age, r.region, r.ip_address, r.created_at.strftime('%d.%m.%Y %H:%M')]
        answers_dict = {a.question_number: a.answer for a in r.answers.all()}
        for q in QUESTIONS:
            if q['type'] != 'consent':
                val = answers_dict.get(q['number'], '')
                try:
                    parsed = json.loads(val)
                    if isinstance(parsed, list):
                        val = '; '.join(parsed)
                except (json.JSONDecodeError, TypeError):
                    pass
                row.append(val)
        writer.writerow(row)

    return response


@staff_member_required
def export_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Respondentlar"

    headers = ['ID', 'Ism', 'Yosh', 'Viloyat', 'IP', 'Sana']
    for q in QUESTIONS:
        if q['type'] != 'consent':
            headers.append(f"Savol {q['number']}")
    ws.append(headers)

    bold = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).font = bold

    respondents = Response.objects.all().prefetch_related('answers')
    for r in respondents:
        row = [r.id, r.name or '', r.age, r.region, r.ip_address, r.created_at.strftime('%d.%m.%Y %H:%M')]
        answers_dict = {a.question_number: a.answer for a in r.answers.all()}
        for q in QUESTIONS:
            if q['type'] != 'consent':
                val = answers_dict.get(q['number'], '')
                try:
                    parsed = json.loads(val)
                    if isinstance(parsed, list):
                        val = '; '.join(parsed)
                except (json.JSONDecodeError, TypeError):
                    pass
                row.append(val)
        ws.append(row)

    for col_idx in range(1, len(headers) + 1):
        max_length = max((len(str(ws.cell(row=row_idx, column=col_idx).value or '')) for row_idx in range(1, ws.max_row + 1)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 60)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="respondents.xlsx"'
    return response
